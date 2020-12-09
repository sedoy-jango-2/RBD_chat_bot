import telebot
from telebot import types

import os
from dotenv import load_dotenv
from datetime import datetime
from dateutil.relativedelta import relativedelta

import pymysql.cursors
import csv
from openpyxl import Workbook


user_dict = {}
load_dotenv()
TOKEN = os.environ.get('BOT_TOKEN')
bot = telebot.TeleBot(TOKEN)
CLUB_ID = 1
connection = pymysql.connect(host='localhost',
                             user='root',
                             password='password',
                             db='fitness',
                             port=3306,
                             cursorclass=pymysql.cursors.DictCursor)


class User:
    def __init__(self):
        self.id = None
        self.surname = ''
        self.name = ''
        self.phone = ''
        self.email = ''
        self.birthdate = ''


class Subscriptions:
    def __init__(self):
        self.subscriptions = {}

    def insert_in_subs(self, id, ClientID, StartDate, EndDate):
        insertion = {
            'ClientID': ClientID,
            'StartDate': StartDate,
            'EndDate': EndDate
        }
        self.subscriptions[id] = insertion


subs_data = Subscriptions()


@bot.message_handler(commands=['start', 'help'])
def send_welcome(message):
    help_message = 'Добрый день!\n\n' + \
                   'Этот бот является терминалом базы данных фитнесклуба Aquastar.\n\n' + \
                   'Пожалуйста, выберите одно из предложенных действий:\n' + \
                   ' - "Контроль абонементов"\n' + \
                   ' - "Контроль посещений"\n'

    bot.send_message(message.chat.id, help_message)
    markup = types.ReplyKeyboardMarkup(one_time_keyboard=True)
    markup.row('Контроль абонементов')
    markup.row('Контроль посещений')
    msg = bot.reply_to(
        message,
        'Что вы хотите сделать?',
        reply_markup=markup
    )
    bot.register_next_step_handler(
        msg,
        control_first_step
    )


def control_first_step(message):
    try:
        if message.text == 'Контроль абонементов':
            markup = types.ReplyKeyboardMarkup(one_time_keyboard=True)
            markup.row('Оформить абонемент')
            markup.row('Аннулировать абонемент')
            markup.row('Выгрузить таблицу клиентов')
            bot.register_next_step_handler(
                bot.reply_to(
                    message,
                    'Выберете следующий пункт меню: ',
                    reply_markup=markup
                ),
                subscriptions_control
            )
        elif message.text == 'Контроль посещений':
            markup = types.ReplyKeyboardMarkup(one_time_keyboard=True)
            markup.row('Впустить клиента')
            markup.row('Выпустить клиента')
            markup.row('Выгрузить таблицу посещений')
            bot.register_next_step_handler(
                bot.reply_to(
                    message,
                    'Выберете следующий пункт меню: ',
                    reply_markup=markup
                ),
                attendance_control
            )

            return
        else:
            return
    except Exception as e:
        bot.reply_to(message, 'oooops')


def subscriptions_control(message):
    if message.text == 'Оформить абонемент':
        markup = types.ReplyKeyboardRemove(selective=False)
        phone_number = bot.reply_to(
            message,
            'Введите номер телефона: ',
            reply_markup=markup
        )
        bot.register_next_step_handler(phone_number, check_phone_number_add)
    elif message.text == 'Аннулировать абонемент':
        markup = types.ReplyKeyboardRemove(selective=False)
        phone_number = bot.reply_to(
            message,
            'Введите номер телефона: ',
            reply_markup=markup
        )
        bot.register_next_step_handler(phone_number, check_phone_number_delete)
    elif message.text == 'Выгрузить таблицу клиентов':
        sql = 'SELECT * FROM CLIENTS'
        with connection.cursor() as cursor: cursor.execute(sql)

        rows = cursor.fetchall()
        headers = [col[0] for col in cursor.description]
        workbook = Workbook()
        sheet = workbook.active

        for i in range(len(headers)):
            sheet.cell(row=1, column=i+1).value = headers[i]
        for row in range(len(rows)):
            for i in range(len(headers)):
                sheet.cell(row=row+2, column=i+1).value = rows[row][headers[i]]
        workbook.save(filename="CLIENTS.xlsx")
        bot.send_document(message.chat.id, open('CLIENTS.xlsx', 'rb'))
        bot.send_message(message.chat.id, 'Для входа в главное меню, введите команду /start')
        return


def check_phone_number_add(message):
    sql = 'SELECT id, surname, name FROM CLIENTS WHERE Phone LIKE "' + message.text + '"'
    with connection.cursor() as cursor: cursor.execute(sql)
    rows = cursor.fetchall()

    user = User()
    user_dict[message.chat.id] = user
    user_dict[message.chat.id].phone = message.text

    if cursor.rowcount > 0:
        new_message = 'На этот номер уже оформлен абонемент.\n' + \
                      'Абонемент оформлен на имя: ' + rows[0]['name'] + ' ' + rows[0]['surname']
        bot.send_message(message.chat.id, new_message)

        user_dict[message.chat.id].id = rows[0]['id']

        markup = types.ReplyKeyboardMarkup(one_time_keyboard=True)
        markup.row('Оформить новый абонемент на это имя')
        markup.row('Выйти в главное меню')
        bot.register_next_step_handler(
            bot.reply_to(
                message,
                'Выберете следующий пункт меню: ',
                reply_markup=markup
            ),
            make_new_subscription
        )
    else:
        new_message = 'На этот номер не оформлен ни один абонемент.\n' + \
                      'Хотите добавить нового пользователя?'
        bot.send_message(message.chat.id, new_message)
        markup = types.ReplyKeyboardMarkup(one_time_keyboard=True)
        markup.row('Да')
        markup.row('Нет')
        bot.register_next_step_handler(
            bot.reply_to(
                message,
                'Выберете пункт меню: ',
                reply_markup=markup
            ),
            add_username
        )


def check_phone_number_delete(message):
    sql = 'SELECT * FROM clients ' + \
          'INNER JOIN subscriptions ' + \
          'ON clients.id = subscriptions.ClientID '+ \
          'WHERE EndDate > SYSDATE() AND ' + \
          'Phone LIKE "' + message.text + '"'
    with connection.cursor() as cursor: cursor.execute(sql)
    rows = cursor.fetchall()
    user = User()
    user_dict[message.chat.id] = user
    user_dict[message.chat.id].phone = message.text

    if cursor.rowcount > 0:
        new_message = 'На этот номер оформлены слудующие абонементы:\n'
        bot.send_message(message.chat.id, new_message)
        markup = types.ReplyKeyboardMarkup(one_time_keyboard=True)
        for row in rows:

            subs_message = 'На имя: ' + row['Name'] + ' ' + \
                           row['Surname'] + ' до ' + row['EndDate'].strftime("%Y-%m-%d")
            bot.send_message(message.chat.id, subs_message)
            markup.row('id: ' + str(row['subscriptions.id']) + '. End date: ' + row['EndDate'].strftime("%Y-%m-%d"))
            subs_data.insert_in_subs(row['subscriptions.id'], row['ClientID'], row['StartDate'], row['EndDate'])
        user_dict[message.chat.id].id = rows[0]['id']
        markup.row('Никакой')
        bot.register_next_step_handler(
            bot.reply_to(
                message,
                'Какой абонемент вы хотите анулировать?',
                reply_markup=markup
            ),
            delete_subscription
        )
    else:
        new_message = 'На этот номер не оформлен ни один абонемент.\n' + \
                      'Хотите добавить нового пользователя?'
        bot.send_message(message.chat.id, new_message)
        markup = types.ReplyKeyboardMarkup(one_time_keyboard=True)
        markup.row('Да')
        markup.row('Нет')
        bot.register_next_step_handler(
            bot.send_message(
                message.chat.id,
                'Выберете пункт меню: ',
                reply_markup=markup
            ),
            add_username
        )


def add_username(message):
    if message.text == 'Да':
        markup = types.ReplyKeyboardRemove(selective=False)
        bot.register_next_step_handler(
            bot.send_message(
                message.chat.id,
                'Введите ФИО: ',
                reply_markup=markup
            ),
            add_birthdate
        )
    else:
        bot.send_message(message.chat.id, 'Для входа в главное меню, введите команду /start')
        return


def add_birthdate(message):
    surname, name, patronymic = message.text.split(' ')

    user_dict[message.chat.id].surname = surname
    user_dict[message.chat.id].name = name

    markup = types.ReplyKeyboardRemove(selective=False)

    bot.register_next_step_handler(
        bot.send_message(
            message.chat.id,
            'Введите дату в формате yyyy-mm-dd: ',
            reply_markup=markup
        ),
        add_email
    )


def add_email(message):
    user_dict[message.chat.id].birthdate = message.text
    markup = types.ReplyKeyboardRemove(selective=False)

    bot.register_next_step_handler(
        bot.send_message(
            message.chat.id,
            'Введите свой email: ',
            reply_markup=markup
        ),
        add_user
    )


def add_user(message):
    user_dict[message.chat.id].email = message.text
    sql = 'INSERT INTO Clients (Surname,Name,Phone,Email,BirthDate) ' + \
          'VALUES ("' + user_dict[message.chat.id].surname + '","' + \
          user_dict[message.chat.id].name + '","' + \
          user_dict[message.chat.id].phone + '","' + \
          user_dict[message.chat.id].email + '","' + \
          user_dict[message.chat.id].birthdate + '")'
    with connection.cursor() as cursor: cursor.execute(sql)
    connection.commit()
    markup = types.ReplyKeyboardMarkup(one_time_keyboard=True)

    sql = 'SELECT id FROM CLIENTS WHERE Phone LIKE "' + \
          user_dict[message.chat.id].phone + '"'
    with connection.cursor() as cursor:
        cursor.execute(sql)
    rows = cursor.fetchall()
    user_dict[message.chat.id].id = rows[0]['id']

    markup.row('Оформить новый абонемент на это имя')
    markup.row('Выйти в главное меню')
    bot.register_next_step_handler(
        bot.send_message(
            message.chat.id,
            'Выберете следующий пункт меню: ',
            reply_markup=markup
        ),
        make_new_subscription
    )


def make_new_subscription(message):
    if message.text == 'Выйти в главное меню':
        bot.send_message(message.chat.id, 'Для входа в главное меню, введите команду /start')
        return
    elif message.text == 'Оформить новый абонемент на это имя':
        markup = types.ReplyKeyboardRemove(selective=False)
        month_period = bot.send_message(
            message.chat.id,
            'На сколько месяцев вы хотите продлить абонемент?',
            reply_markup=markup
        )

        bot.register_next_step_handler(
            month_period,
            make_new_subscription_query
        )
    else:
        return


def make_new_subscription_query(message):
    now_date = datetime.now()
    now_date_str = now_date.strftime("%Y-%m-%d")
    new_date = now_date + relativedelta(months=int(message.text))
    new_date_str = new_date.strftime("%Y-%m-%d")
    sql = 'INSERT INTO Subscriptions(ClientID, StartDate, EndDate) ' + \
          'VALUES(' + str(user_dict[message.chat.id].id) + ', "' + now_date_str + '", "' + new_date_str + '")'
    with connection.cursor() as cursor: cursor.execute(sql)
    connection.commit()

    sql = 'SELECT * FROM Subscriptions WHERE ClientID LIKE ' + str(user_dict[message.chat.id].id) + \
        ' AND StartDate LIKE "' + datetime.today().strftime("%Y-%m-%d") + '"'
    with connection.cursor() as cursor: cursor.execute(sql)
    rows = cursor.fetchall()

    bot.send_message(message.chat.id, 'Вы оформили абонемент на ' + str(message.text) + ' месяцев')
    bot.send_message(message.chat.id, 'Номер вашего абонемента ' + rows[0]['id'])
    bot.send_message(message.chat.id, 'Для входа в главное меню, введите команду /start')
    user_dict.pop(message.chat.id)
    return


def delete_subscription(message):
    if message.text == 'Никакой':
        bot.send_message(message.chat.id, 'Для входа в главное меню, введите команду /start')
        return
    subs_id = int(message.text[message.text.find('id: ') + 4:message.text.find('. End date: ')])
    client_id = subs_data.subscriptions[subs_id]['ClientID']
    start_date = subs_data.subscriptions[subs_id]['StartDate']
    end_date = subs_data.subscriptions[subs_id]['EndDate']
    now_date = datetime.today()
    now_date -= relativedelta(days=1)
    sql = 'UPDATE subscriptions SET EndDate="' + now_date.strftime("%Y-%m-%d") + \
          '" where subscriptions.id like ' + str(subs_id)
    with connection.cursor() as cursor: cursor.execute(sql)
    connection.commit()
    bot.send_message(message.chat.id, 'Абонемент с номером ' + str(subs_id) + ' анулирован')
    bot.send_message(message.chat.id, 'Для выхода в главное меню, введите команду /start')
    return


def attendance_control(message):

    if message.text == 'Впустить клиента':
        markup = types.ReplyKeyboardRemove(selective=False)
        phone_number = bot.send_message(
            message.chat.id,
            'Введите номер абонемента: ',
            reply_markup=markup
        )
        bot.register_next_step_handler(phone_number, check_client_to_get_in)
        return
    elif message.text == 'Выпустить клиента':
        markup = types.ReplyKeyboardRemove(selective=False)
        phone_number = bot.send_message(
            message.chat.id,
            'Введите номер абонемента: ',
            reply_markup=markup
        )
        bot.register_next_step_handler(phone_number, check_client_to_get_out)
        return
    elif message.text == 'Выгрузить таблицу посещений':
        sql = \
            'SELECT Name, Surname, Phone, SubscriptionID, EntryTime, ExitTime, City, Street from \
        (SELECT EntryTime, ExitTime, City, Street, ClientID, SubscriptionID \
        from (select SubscriptionID, EntryTime, ExitTime, City, Street \
        from attendance INNER JOIN clubs ON clubs.id=ClubID) as T \
        inner join subscriptions on SubscriptionID=subscriptions.id) as TT \
        inner join clients on ClientID=clients.id'
        with connection.cursor() as cursor: cursor.execute(sql)
        rows = cursor.fetchall()
        headers = [col[0] for col in cursor.description]
        workbook = Workbook()
        sheet = workbook.active

        for i in range(len(headers)):
            sheet.cell(row=1, column=i + 1).value = headers[i]
        for row in range(len(rows)):
            for i in range(len(headers)):
                sheet.cell(row=row + 2, column=i + 1).value = rows[row][headers[i]]
        workbook.save(filename="ATTENDANCE.xlsx")
        bot.send_document(message.chat.id, open('ATTENDANCE.xlsx', 'rb'))
        bot.send_message(message.chat.id, 'Для выхода в главное меню, введите команду /start')
        return


def check_client_to_get_in(message):
    sql = \
        'SELECT Name, Surname, Phone, SubscriptionID, EntryTime, ExitTime, City, Street from \
    (SELECT EntryTime, ExitTime, City, Street, ClientID, SubscriptionID \
    from (select SubscriptionID, EntryTime, ExitTime, City, Street \
    from attendance INNER JOIN clubs ON clubs.id=ClubID) as T \
    inner join subscriptions on SubscriptionID=subscriptions.id) as TT \
    inner join clients on ClientID=clients.id \
    where ExitTime like NULL and SubscriptionID like ' + message.text

    with connection.cursor() as cursor: cursor.execute(sql)
    if cursor.rowcount > 0:
        bot.send_message(message.chat.id, 'По вашему абонементу кто-то уже занимается')
        markup = types.ReplyKeyboardMarkup(one_time_keyboard=True)
        markup.row('Оформить новый абонемент на это имя')
        markup.row('Выйти в главное меню')
        rows = cursor.fetchall()
        user = User()
        user_dict[message.chat.id] = user
        user_dict[message.chat.id].phone = rows[0]['Phone']
        bot.register_next_step_handler(
            bot.send_message(
                message.chat.id,
                'Выберете следующий пункт меню: ',
                reply_markup=markup
            ),
            make_new_subscription
        )
    else:
        sql = \
            'select name, surname from clients inner join ' + \
            'subscriptions on clients.id = ClientID ' + \
            'where subscriptions.id like ' + message.text

        with connection.cursor() as cursor:
            cursor.execute(sql)

        rows = cursor.fetchall()
        new_message = 'Добро пожаловать, ' + rows[0]['name'] + ' ' + rows[0]['surname']
        bot.send_message(message.chat.id, new_message)
        sql = \
            'INSERT INTO Attendance (SubscriptionID,ClubID,EntryTime,ExitTime) \
            VALUES (' + message.text + ',' + str(CLUB_ID) + ',"' + datetime.today().strftime("%Y-%m-%d") + '",NULL)'

        with connection.cursor() as cursor:
            cursor.execute(sql)
        connection.commit()
        return
    return


def check_client_to_get_out(message):
    sql = \
        'SELECT a_id, Name, Surname, Phone, SubscriptionID, EntryTime, ExitTime, City, Street from \
    (SELECT a_id, EntryTime, ExitTime, City, Street, ClientID, SubscriptionID \
    from (select attendance.id a_id, SubscriptionID, EntryTime, ExitTime, City, Street \
    from attendance INNER JOIN clubs ON clubs.id=ClubID) as T \
    inner join subscriptions on SubscriptionID=subscriptions.id) as TT \
    inner join clients on ClientID=clients.id \
    where ExitTime IS NULL and SubscriptionID like "' + message.text + '"'

    with connection.cursor() as cursor: cursor.execute(sql)

    if cursor.rowcount > 0:
        rows = cursor.fetchall()
        bot.send_message(message.chat.id, 'Спасибо, что посетили наш фитнесклуб, '\
                         + rows[0]['Name'] + ' ' + rows[0]['Surname'])

        sql = 'UPDATE Attendance SET ExitTime="' + datetime.now().strftime("%Y-%m-%d %H:%M:%S") + \
              '" where Attendance.id like ' + str(rows[0]['a_id'])

        with connection.cursor() as cursor:
            cursor.execute(sql)
        connection.commit()
    else:
        new_message = 'Ошибка, попробуйте еще раз'
        bot.send_message(message.chat.id, new_message)

        markup = types.ReplyKeyboardRemove(selective=False)
        phone_number = bot.send_message(
            message.chat.id,
            'Введите номер абонемента: ',
            reply_markup=markup
        )
        bot.register_next_step_handler(phone_number, check_client_to_get_out)
        return


# Кеширование
bot.enable_save_next_step_handlers(delay=2)
bot.load_next_step_handlers()

bot.polling()
